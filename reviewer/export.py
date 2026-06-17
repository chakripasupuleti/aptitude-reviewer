import re
from pathlib import Path

import pandas as pd

# Characters that Excel's XML format cannot represent
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]")


def _sanitize(value):
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub("", value)
    return value


def export_reviewed(df: pd.DataFrame, output_path: str) -> None:
    path = Path(output_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        df.to_csv(path, index=False)
        return

    if suffix in {".xlsx", ".xlsm"}:
        df = df.map(_sanitize)
        df.to_excel(path, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            workbook = load_workbook(path)
            worksheet = workbook.active
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells[:50])
                width = min(max(max_length + 2, 12), 60)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width
            workbook.save(path)
        except Exception:
            # Formatting failure should not block the reviewed output.
            pass
        return

    raise ValueError("Unsupported output file type. Use .csv or .xlsx")
